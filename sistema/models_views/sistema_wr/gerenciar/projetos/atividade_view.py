from sistema import app, requires_roles, db
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sistema.models_views.upload_arquivo.upload_arquivo_view import upload_arquivo
from sistema.models_views.sistema_wr.gerenciar.projetos.atividade_model import AtividadeModel, AtividadeAnexoModel
from sistema.models_views.sistema_wr.gerenciar.projetos.atividade_prioridade_model import PrioridadeAtividadeModel
from sistema.models_views.sistema_wr.gerenciar.projetos.atividade_andamento_model import AndamentoAtividadeModel
from sistema.models_views.sistema_wr.gerenciar.projetos.atividade_solicitacao_model import SolicitacaoAtividadeModel
from sistema.models_views.sistema_wr.gerenciar.projetos.projeto_model import ProjetoModel
from sistema.models_views.sistema_wr.autenticacao.usuario_model import UsuarioModel
from sistema.models_views.sistema_wr.configuracoes.tags.tags_model import TagModel
from sistema.models_views.sistema_wr.configuracoes.categorias.categoria_model import CategoriaModel
from sistema.models_views.sistema_wr.configuracoes.email_atividades.email_atividade_model import EmailAtividadeModel
from sistema.models_views.sistema_wr.configuracoes.area_atividades.area_atividade_model import AreaModel
from sistema._utilitarios import *
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from sqlalchemy.orm import joinedload


# Configurações de upload
ALLOWED_EXTENSIONS = {
    'image': {'png', 'jpg', 'jpeg', 'gif', 'webp'},
    'document': {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'rtf'},
    'compressed': {'zip', 'rar', '7z'},
    'other': {'csv'}
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def _extrair_tag_ids_do_request(form_or_args, key="tag_id"):
    """
    Extrai IDs de tags aceitando:
    - múltiplos valores (checkboxes): getlist('tag_id')
    - valor único (select antigo): get('tag_id')
    Retorna lista de strings (sem vazios).
    """
    ids = []
    try:
        ids = form_or_args.getlist(key) or []
    except Exception:
        ids = []
    
    if not ids:
        unico = form_or_args.get(key)
        if unico:
            ids = [unico]
    
    return [str(x).strip() for x in ids if  x is not None and str(x).strip()]

def _processar_tags_selecionadas(ids_str, campos_erros, dados_corretos, key_erro="tag_id"):
    """
    Valida IDs de tags, busca tags ativas e retorna lista de TagModel.
    Preenche dados_corretos['tag_id'] (lista de strings) para re-render do form.
    
    Args:
        ids_str: lista de IDs como strings (ex: ['1', '3', '5'])
        campos_erros: dict para adicionar mensagens de erro
        dados_corretos: dict para manter valores no form após erro
        key_erro: chave do erro no dict (padrão 'tag_id')
    
    Returns:
        Lista de TagModel (vazia se sem tags ou erro)
    """

    dados_corretos["tag_id"] = ids_str

    if not ids_str:
        return []

    #Converter para ints e remover duplicados

    try:
        ids_int = sorted({int(x) for x in ids_str})
    except(ValueError, TypeError):
        campos_erros[key_erro] = "IDs de tags inválidos"
        return []
    
    tags = TagModel.query.filter(
        TagModel.id.in_(ids_int),
        TagModel.deletado == False
    ).all()

    if len(tags) != len(ids_int):
        campos_erros[key_erro] = "Uma ou mais tags selecionadas não existem ou foram removidas"
        return []
    
    return tags


def allowed_file(filename):
    """Verifica se o arquivo é permitido baseado nas extensões configuradas"""
    if '.' not in filename:
        return False
    
    ext = filename.rsplit('.', 1)[1].lower()
    
    # Verifica em todas as categorias
    for extensions in ALLOWED_EXTENSIONS.values():
        if ext in extensions:
            return True
    return False


def get_file_type(filename):
    """Determina o tipo do arquivo baseado na extensão"""
    if '.' not in filename:
        return 'other'
    
    ext = filename.rsplit('.', 1)[1].lower()
    
    for file_type, extensions in ALLOWED_EXTENSIONS.items():
        if ext in extensions:
            return file_type
    return 'other'


def get_cor_prioridade(prioridade_id):
    """Retorna a cor do badge baseada na prioridade"""
    cores = {
        1: 'danger',   # Alta
        2: 'warning',  # Média
        3: 'primary',  # Baixa
        4: 'secondary' # Muito baixa
    }
    return cores.get(prioridade_id, 'secondary')


def get_cor_situacao(situacao_id):
    """Retorna a cor do badge baseada na situação"""
    cores = {
        1: 'secondary', # Não iniciada
        2: 'warning',   # Em andamento
        3: 'info',      # Pausada
        4: 'success',   # Concluída
        5: 'danger',    # Cancelada
        6: 'orange',    # Atrasada
        7: 'primary',   # Em Análise
        8: 'dark'       # Rejeitada
    }
    return cores.get(situacao_id, 'secondary')


@app.route("/gerenciar/atividades/anexo/download/<int:anexo_id>")
@login_required
@requires_roles
def atividade_download_anexo(anexo_id):
    anexo = AtividadeAnexoModel.query.filter(
        AtividadeAnexoModel.id == anexo_id,
        AtividadeAnexoModel.deletado == False
    ).first()
    
    if not anexo:
        flash(("Anexo não encontrado!", "warning"))
        referrer = request.referrer
        if referrer and 'atividade' in referrer:
            return redirect(referrer)
        return redirect(url_for("atividades_listar"))
    
    # Encontra o caminho correto do arquivo
    def encontrar_arquivo():
        caminhos_teste = [
            anexo.caminho_arquivo,
            os.path.join(os.getcwd(), anexo.caminho_arquivo),
            anexo.caminho_arquivo if os.path.isabs(anexo.caminho_arquivo) else None
        ]
        
        for caminho in caminhos_teste:
            if caminho and os.path.exists(caminho):
                # Garante que e retornado o caminho absoluto para o send_file
                return os.path.abspath(caminho)
        
        return None
    
    caminho_arquivo = encontrar_arquivo()
    
    if not caminho_arquivo:
        flash(("Arquivo não encontrado no servidor!", "warning"))
        return redirect(url_for("atividade_visualizar", atividade_id=anexo.atividade_id))
    
    try:
        # Faz o download com o caminho absoluto correto
        return send_file(
            caminho_arquivo,
            as_attachment=True,
            download_name=anexo.nome_original,
            mimetype=anexo.mime_type if anexo.mime_type else 'application/octet-stream'
        )
    except Exception as e:
        return redirect(url_for("atividade_visualizar", atividade_id=anexo.atividade_id))
    
    
@app.route("/gerenciar/atividades/visualizar/<int:atividade_id>")
@login_required
@requires_roles
def atividade_visualizar(atividade_id):
    atividade = AtividadeModel.obter_atividade_por_id(atividade_id)
    
    if not atividade:
        flash(("Atividade não encontrada!", "warning"))
        return redirect(url_for("atividades_listar"))
    
    # Buscar anexos da atividade
    anexos = atividade.anexos.filter(AtividadeAnexoModel.deletado == False).all()
    
    # Calcular estatísticas
    percentual_horas = atividade.percentual_horas
    percentual_horas_visual = min(percentual_horas, 100)  # Para a barra de progresso
    esta_atrasada = atividade.esta_atrasada
    
    # Informações do valor
    valor_real = atividade.valor_atividade_100 / 100 if atividade.valor_atividade_100 else 0
    
    # Formatação de datas
    data_prazo_formatada = None
    if atividade.data_prazo_conclusao:
        data_prazo_formatada = atividade.data_prazo_conclusao.strftime('%d/%m/%Y')
    
    data_cadastro_formatada = atividade.data_cadastro.strftime('%d/%m/%Y às %H:%M')
    data_alteracao_formatada = atividade.data_alteracao.strftime('%d/%m/%Y às %H:%M')
    
    # Status da atividade para cores dos badges
    status_cores = {
        'prioridade': get_cor_prioridade(atividade.prioridade_id),
        'situacao': get_cor_situacao(atividade.situacao_id)
    }
    
    return render_template(
        "sistema_wr/gerenciar/projetos/atividade_visualizar.html",
        atividade=atividade,
        anexos=anexos,
        percentual_horas=percentual_horas,
        percentual_horas_visual=percentual_horas_visual,
        esta_atrasada=esta_atrasada,
        valor_real=valor_real,
        data_prazo_formatada=data_prazo_formatada,
        data_cadastro_formatada=data_cadastro_formatada,
        data_alteracao_formatada=data_alteracao_formatada,
        status_cores=status_cores
    )
    
    
@app.route("/gerenciar/atividades")
@app.route("/gerenciar/atividades/<int:projeto_id>")
@login_required
@requires_roles
def atividades_listar(projeto_id=None):
    filtro_data_conclusao_inicio = request.args.get('data_conclusao_inicio')
    filtro_data_conclusao_fim = request.args.get('data_conclusao_fim')
    
    # Modificado para suportar múltiplos valores
    filtros_projeto = request.args.getlist('projeto_id')
    filtros_prioridade = request.args.getlist('prioridade_id')
    filtros_situacao = request.args.getlist('situacao_id')
    filtros_responsavel = request.args.getlist('responsavel_id')
    filtros_solicitante = request.args.getlist('solicitante_id')
    filtros_supervisor = request.args.getlist('supervisor_id')
    filtros_categoria = request.args.getlist('categoria_id')
    filtros_area = request.args.getlist('area_id')
    


    filtros_tag = [t for t in request.args.getlist('tag_id') if t and str(t).strip()]

    filtro_titulo = request.args.get('titulo')
    
    # Se veio projeto_id pela URL, adicionar aos filtros
    if projeto_id:
        if str(projeto_id) not in filtros_projeto:
            filtros_projeto.append(str(projeto_id))
    
    # Remover valores vazios dos filtros múltiplos
    filtros_projeto = [p for p in filtros_projeto if p and p.strip()]
    filtros_prioridade = [p for p in filtros_prioridade if p and p.strip()]
    filtros_situacao = [s for s in filtros_situacao if s and s.strip()]
    filtros_responsavel = [r for r in filtros_responsavel if r and r.strip()]
    filtros_solicitante = [s for s in filtros_solicitante if s and s.strip()]
    filtros_supervisor = [s for s in filtros_supervisor if s and s.strip()]
    filtros_categoria = [c for c in filtros_categoria if c and c.strip()]
    filtros_area = [a for a in filtros_area if a and a.strip()]

    atividades = AtividadeModel.query.filter(
        AtividadeModel.deletado == False
    ).options(
        joinedload(AtividadeModel.tags)
    )


    if filtro_data_conclusao_inicio:
        try:
            data_inicio_obj = datetime.strptime(filtro_data_conclusao_inicio, '%Y-%m-%d').date()
            atividades = atividades.filter(AtividadeModel.data_prazo_conclusao >= data_inicio_obj)
        except ValueError:
            pass

    if filtro_data_conclusao_fim:
        try:
            data_fim_obj = datetime.strptime(filtro_data_conclusao_fim, '%Y-%m-%d').date()
            atividades = atividades.filter(AtividadeModel.data_prazo_conclusao <= data_fim_obj)
        except ValueError:
            pass
    
    # Aplicar filtros múltiplos usando IN
    if filtros_projeto:
        atividades = atividades.filter(AtividadeModel.projeto_id.in_(filtros_projeto))
    
    if filtros_prioridade:
        atividades = atividades.filter(AtividadeModel.prioridade_id.in_(filtros_prioridade))
    
    if filtros_situacao:
        atividades = atividades.filter(AtividadeModel.situacao_id.in_(filtros_situacao))
    
    if filtros_responsavel:
        atividades = atividades.filter(AtividadeModel.desenvolvedor_id.in_(filtros_responsavel))
        
    if filtros_solicitante:
        atividades = atividades.filter(AtividadeModel.usuario_solicitante_id.in_(filtros_solicitante))
    
    if filtros_supervisor:
        atividades = atividades.filter(AtividadeModel.supervisor_id.in_(filtros_supervisor))

    if filtros_tag:
        try:
            tag_id_int = [int(x) for x in filtros_tag]
            atividades = atividades.filter(
                AtividadeModel.tags.any(TagModel.id.in_(tag_id_int))
            )
        except(ValueError, TypeError):
            pass
    
    if filtro_titulo:
        atividades = atividades.filter(AtividadeModel.titulo.ilike(f"%{filtro_titulo}%"))
    
    if filtros_categoria:
        atividades = atividades.filter(AtividadeModel.categoria_id.in_(filtros_categoria))
    
    if filtros_area:
        atividades = atividades.filter(AtividadeModel.area_id.in_(filtros_area))
    
    atividades = atividades.order_by(
        AtividadeModel.id.desc(),
        AtividadeModel.data_cadastro.desc()
    ).all()
    
    categorias = CategoriaModel.listar_categorias_ativas()
    areas = AreaModel.listar_areas_ativas()
    projetos = ProjetoModel.obter_projetos_asc_nome()
    prioridades = PrioridadeAtividadeModel.listar_prioridades_ativas()
    situacoes = AndamentoAtividadeModel.listar_andamentos_ativos()
    usuarios = UsuarioModel.obter_usuarios_asc_nome()
    
    return render_template(
        "sistema_wr/gerenciar/projetos/atividades_listar.html",
        atividades=atividades,
        projetos=projetos,
        prioridades=prioridades,
        situacoes=situacoes,
        usuarios=usuarios,
        tags=TagModel.listar_tags_ativas(),
        categorias=categorias,
        areas=areas,
        filtros={
            'data_conclusao_inicio': filtro_data_conclusao_inicio,
            'data_conclusao_fim': filtro_data_conclusao_fim,
            'projeto_id': filtros_projeto,
            'prioridade_id': filtros_prioridade,
            'situacao_id': filtros_situacao,
            'responsavel_id': filtros_responsavel,
            'solicitante_id': filtros_solicitante,
            'supervisor_id': filtros_supervisor,
            'tag_id': filtros_tag,
            'titulo': filtro_titulo,
            'categoria_id': filtros_categoria,
            'area_id': filtros_area
        }
    )


@app.route("/gerenciar/atividades/exportar")
@login_required
@requires_roles
def atividades_exportar():
    """Exporta atividades filtradas para Excel com design moderno e resumo"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Obter filtros
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    filtros_projeto = [p for p in request.args.getlist('projeto_id') if p and p.strip()]
    filtros_situacao = [s for s in request.args.getlist('situacao_id') if s and s.strip()]
    filtros_categoria = [c for c in request.args.getlist('categoria_id') if c and c.strip()]
    filtros_area = [a for a in request.args.getlist('area_id') if a and a.strip()]
    
    # Query base
    atividades = AtividadeModel.query.filter(AtividadeModel.deletado == False)
    
    # Aplicar filtros
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            atividades = atividades.filter(AtividadeModel.data_prazo_conclusao >= data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            atividades = atividades.filter(AtividadeModel.data_prazo_conclusao <= data_fim_obj)
        except ValueError:
            pass
    
    if filtros_projeto:
        atividades = atividades.filter(AtividadeModel.projeto_id.in_(filtros_projeto))
    
    if filtros_situacao:
        atividades = atividades.filter(AtividadeModel.situacao_id.in_(filtros_situacao))
    
    if filtros_categoria:
        atividades = atividades.filter(AtividadeModel.categoria_id.in_(filtros_categoria))
    
    if filtros_area:
        atividades = atividades.filter(AtividadeModel.area_id.in_(filtros_area))
    
    atividades = atividades.order_by(AtividadeModel.id.desc()).all()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Atividades"
    
    # ========== PALETA DE CORES ==========
    AZUL_PRIMARIO = "1a73e8"
    AZUL_ESCURO = "0d47a1"
    VERDE_SUCESSO = "34a853"
    AMARELO_ALERTA = "fbbc04"
    VERMELHO_PERIGO = "ea4335"
    CINZA_CLARO = "f8f9fa"
    CINZA_MEDIO = "e9ecef"
    BRANCO = "ffffff"
    PRETO = "212529"
    
    # ========== ESTILOS ==========
    thin_border = Border(
        left=Side(style='thin', color=CINZA_MEDIO),
        right=Side(style='thin', color=CINZA_MEDIO),
        top=Side(style='thin', color=CINZA_MEDIO),
        bottom=Side(style='thin', color=CINZA_MEDIO)
    )
    
    # ========== CABEÇALHO DO RELATÓRIO ==========
    num_colunas = 12  # Total de colunas da tabela
    ws.merge_cells(f'A1:{get_column_letter(num_colunas)}1')
    header_cell = ws['A1']
    header_cell.value = "RELATORIO DE ATIVIDADES WR"
    header_cell.font = Font(bold=True, size=20, color=BRANCO)
    header_cell.fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Subtítulo
    ws.merge_cells(f'A2:{get_column_letter(num_colunas)}2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M')} | Por: {current_user.nome} {current_user.sobrenome}"
    subtitle_cell.font = Font(italic=True, size=10, color=BRANCO)
    subtitle_cell.fill = PatternFill(start_color=AZUL_PRIMARIO, end_color=AZUL_PRIMARIO, fill_type="solid")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    
    # ========== FILTROS APLICADOS ==========
    filtros_texto = []
    if data_inicio:
        filtros_texto.append(f"Data Início: {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    if data_fim:
        filtros_texto.append(f"Data Fim: {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    if filtros_projeto:
        from sistema.models_views.sistema_wr.gerenciar.projetos.projeto_model import ProjetoModel
        nomes_projetos = [p.nome_projeto for p in ProjetoModel.query.filter(ProjetoModel.id.in_(filtros_projeto)).all()]
        filtros_texto.append(f"Projetos: {', '.join(nomes_projetos[:3])}{'...' if len(nomes_projetos) > 3 else ''}")
    if filtros_situacao:
        nomes_situacoes = [s.nome for s in AndamentoAtividadeModel.query.filter(AndamentoAtividadeModel.id.in_(filtros_situacao)).all()]
        filtros_texto.append(f"Situações: {', '.join(nomes_situacoes)}")
    if filtros_categoria:
        nomes_categorias = [c.nome for c in CategoriaModel.query.filter(CategoriaModel.id.in_(filtros_categoria)).all()]
        filtros_texto.append(f"Categorias: {', '.join(nomes_categorias[:3])}{'...' if len(nomes_categorias) > 3 else ''}")
    if filtros_area:
        from sistema.models_views.sistema_wr.configuracoes.area_atividades.area_atividade_model import AreaModel
        nomes_areas = [a.nome for a in AreaModel.query.filter(AreaModel.id.in_(filtros_area)).all()]
        filtros_texto.append(f"Áreas: {', '.join(nomes_areas)}")
    
    # Linha de filtros (linha 3), só aparece se houver filtros
    current_row = 3
    if filtros_texto:
        ws.merge_cells(f'A{current_row}:{get_column_letter(num_colunas)}{current_row}')
        filtros_cell = ws.cell(row=current_row, column=1)
        filtros_cell.value = "FILTROS: " + "  |  ".join(filtros_texto)
        filtros_cell.font = Font(size=9, color="666666", italic=True)
        filtros_cell.fill = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
        filtros_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1
    
    # ========== CALCULAR MÉTRICAS ==========
    hoje = datetime.now().date()
    total_atividades = len(atividades)
    total_horas_necessarias = sum(a.horas_necessarias or 0 for a in atividades)
    total_horas_utilizadas = sum(a.horas_utilizadas or 0 for a in atividades)
    total_valor = sum((a.valor_atividade_100 or 0) / 100 for a in atividades)
    
    nao_iniciadas = sum(1 for a in atividades if a.situacao_id == 1)
    em_andamento = sum(1 for a in atividades if a.situacao_id == 2)
    pausadas = sum(1 for a in atividades if a.situacao_id == 3)
    concluidas = sum(1 for a in atividades if a.situacao_id == 4)
    canceladas = sum(1 for a in atividades if a.situacao_id == 5)
    atrasadas = sum(1 for a in atividades if a.situacao_id == 6)
    
    # ========== RESUMO EXECUTIVO - CARDS MODERNOS ==========
    ws.merge_cells(f'A{current_row}:{get_column_letter(num_colunas)}{current_row}')
    resumo_title = ws.cell(row=current_row, column=1)
    resumo_title.value = "DASHBOARD DE INDICADORES"
    resumo_title.font = Font(bold=True, size=14, color=AZUL_ESCURO)
    resumo_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1
    
    # KPI Cards - 2 linhas (Label em cima, Valor embaixo)
    kpi_label_row = current_row
    kpi_value_row = current_row + 1
    
    kpi_configs = [
        # (label, value, bg_color, text_color, cols_start, cols_end)
        ("TOTAL", str(total_atividades), "e3f2fd", AZUL_ESCURO, 1, 2),
        ("NÃO INICIADAS", str(nao_iniciadas), "e0e0e0", "6c757d", 3, 4),
        ("EM ANDAMENTO", str(em_andamento), "fff3e0", "e65100", 5, 6),
        ("PAUSADAS", str(pausadas), "e1f5fe", "0277bd", 7, 8),
        ("CONCLUÍDAS", str(concluidas), "e8f5e9", VERDE_SUCESSO, 9, 10),
        ("CANCELADAS", str(canceladas), "ffebee", VERMELHO_PERIGO, 11, 12)
    ]
    
    for label, value, bg_color, text_color, col_start, col_end in kpi_configs:
        # Merge cells para o card
        ws.merge_cells(start_row=kpi_label_row, start_column=col_start, end_row=kpi_label_row, end_column=col_end)
        ws.merge_cells(start_row=kpi_value_row, start_column=col_start, end_row=kpi_value_row, end_column=col_end)
        
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        border_card = Border(
            left=Side(style='thin', color=CINZA_MEDIO),
            right=Side(style='thin', color=CINZA_MEDIO),
            top=Side(style='thin', color=CINZA_MEDIO),
            bottom=Side(style='thin', color=CINZA_MEDIO)
        )
        
        # Label
        label_cell = ws.cell(row=kpi_label_row, column=col_start, value=label)
        label_cell.font = Font(bold=True, size=8, color="666666")
        label_cell.fill = fill
        label_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        label_cell.border = border_card
        
        # Value
        value_cell = ws.cell(row=kpi_value_row, column=col_start, value=value)
        value_cell.font = Font(bold=True, size=16, color=text_color)
        value_cell.fill = fill
        value_cell.alignment = Alignment(horizontal="center", vertical="top")
        value_cell.border = border_card
    
    ws.row_dimensions[kpi_label_row].height = 22
    ws.row_dimensions[kpi_value_row].height = 32
    current_row += 2
    
    # ========== TABELA DE DADOS ==========
    ws.merge_cells(f'A{current_row}:{get_column_letter(num_colunas)}{current_row}')
    table_title = ws.cell(row=current_row, column=1)
    table_title.value = "LISTA DETALHADA DE ATIVIDADES"
    table_title.font = Font(bold=True, size=12, color=AZUL_ESCURO)
    table_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    # Cabeçalhos da tabela (sem Hrs Nec, Hrs Util, Valor, Cadastro)
    headers = [
        "ID", "Título", "Projeto", "Situação", "Prioridade", "Categoria", "Área",
        "Solicitante", "Supervisor", "DEV", "Prazo", "Dias Atraso"
    ]
    
    header_row = current_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, size=10, color=BRANCO)
        cell.fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 28
    
    # Dados
    for idx, atividade in enumerate(atividades):
        row = header_row + 1 + idx
        bg_color = CINZA_CLARO if idx % 2 == 0 else BRANCO
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Cores por situação: 1=Não iniciada, 2=Em andamento, 3=Pausada, 4=Concluída, 5=Cancelada, 6=Atrasada, 7=Em Análise, 8=Rejeitada
        situacao_cores = {
            1: "6c757d",        # Não iniciada - cinza
            2: AMARELO_ALERTA,  # Em andamento - amarelo
            3: "17a2b8",        # Pausada - azul info
            4: VERDE_SUCESSO,   # Concluída - verde
            5: VERMELHO_PERIGO, # Cancelada - vermelho
            6: "ff6b35",        # Atrasada - laranja
            7: "0d6efd",        # Em Análise - azul primário
            8: "343a40"         # Rejeitada - cinza escuro
        }
        situacao_cor = situacao_cores.get(atividade.situacao_id, PRETO)
        
        # Calcular dias de atraso (não aplicável para Concluída=4, Cancelada=5, Atrasada=6)
        dias_atraso = ""
        if atividade.data_prazo_conclusao and atividade.situacao_id not in [4, 5, 6]:
            diff = (hoje - atividade.data_prazo_conclusao).days
            dias_atraso = diff if diff > 0 else 0
        elif atividade.situacao_id in [4, 5, 6]:
            dias_atraso = "-"
        
        dados = [
            atividade.id,
            atividade.titulo[:50] + "..." if len(atividade.titulo or "") > 50 else atividade.titulo,
            atividade.projeto.nome_projeto[:25] + "..." if atividade.projeto and len(atividade.projeto.nome_projeto) > 25 else (atividade.projeto.nome_projeto if atividade.projeto else "-"),
            atividade.situacao.nome if atividade.situacao else "-",
            atividade.prioridade.nome if atividade.prioridade else "-",
            atividade.categoria.nome if atividade.categoria else "-",
            atividade.area.nome if atividade.area else "-",
            f"{atividade.usuario_solicitante.nome} {atividade.usuario_solicitante.sobrenome[0]}." if atividade.usuario_solicitante else "-",
            f"{atividade.supervisor.nome} {atividade.supervisor.sobrenome[0]}." if atividade.supervisor else "-",
            f"{atividade.desenvolvedor.nome} {atividade.desenvolvedor.sobrenome[0]}." if atividade.desenvolvedor else "-",
            atividade.data_prazo_conclusao.strftime('%d/%m/%Y') if atividade.data_prazo_conclusao else "-",
            dias_atraso
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col == 2 else "center", vertical="center")
            cell.font = Font(size=9, color=PRETO)
            
            # Cor especial para situação
            if col == 4:
                cell.font = Font(size=9, color=situacao_cor, bold=True)
            # Cor para dias de atraso
            if col == 12 and isinstance(dias_atraso, int) and dias_atraso > 0:
                cell.font = Font(size=9, color=VERMELHO_PERIGO, bold=True)
        
        ws.row_dimensions[row].height = 20
    
    # ========== APLICAR AUTOFILTRO ==========
    last_data_row = header_row + len(atividades)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_colunas)}{last_data_row}"
    
    # ========== AJUSTAR LARGURA DAS COLUNAS ==========
    column_widths = [6, 50, 28, 14, 12, 15, 15, 18, 18, 18, 12, 11]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # ========== RODAPÉ ==========
    footer_row = last_data_row + 1
    ws.merge_cells(f'A{footer_row}:{get_column_letter(num_colunas)}{footer_row}')
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = f"Total de registros: {total_atividades} | Sistema WR - {datetime.now().year}"
    footer_cell.font = Font(italic=True, size=9, color="666666")
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Salvar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    nome_arquivo = f"relatorio_atividades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route("/gerenciar/atividades/cadastrar", methods=["GET", "POST"])
@app.route("/gerenciar/atividades/cadastrar/<int:projeto_id>", methods=["GET", "POST"])
@app.route("/gerenciar/atividades/cadastrar/<int:projeto_id>/<int:solicitacao_id>", methods=["GET", "POST"])
@login_required
@requires_roles
def atividade_cadastrar(projeto_id=None, solicitacao_id=None):
    validacao_campos_obrigatorios = {}
    validacao_campos_erros = {}
    gravar_banco = True
    categorias = CategoriaModel.listar_categorias_ativas()
    areas = AreaModel.listar_areas_ativas()

    projetos = ProjetoModel.obter_projetos_asc_nome()
    prioridades = PrioridadeAtividadeModel.listar_prioridades_ativas()
    situacoes = AndamentoAtividadeModel.listar_andamentos_ativos()
    usuarios = UsuarioModel.obter_usuarios_asc_nome()
    tags = TagModel.listar_tags_ativas()

    # Dados iniciais (podem vir de uma solicitação aceita)
    dados_corretos = {}
    solicitacao = None
    
    # Se veio de uma solicitação aceita, pré-preencher os dados
    if solicitacao_id:
        solicitacao = SolicitacaoAtividadeModel.obter_solicitacao_por_id(solicitacao_id)
        if solicitacao and solicitacao.situacao_id == 3:  # 3 = Aceita
            dados_corretos = {
                "projetoId": str(solicitacao.projeto_id),
                "titulo": solicitacao.titulo,
                "descricao": solicitacao.descricao or "",
                "prioridadeId": "2",  # Prioridade média como padrão
                "situacaoId": "1",    # Não iniciada como padrão
                "usuarioSolicitanteId": str(solicitacao.usuario_solicitante_id)  # Pré-preencher solicitante
            }
            projeto_id = solicitacao.projeto_id
        else:
            flash(("Solicitação não encontrada ou não aceita!", "warning"))
            return redirect(url_for("solicitacoes_atividade_listar"))

    if request.method == "POST":
        projeto_id = request.form["projetoId"]
        titulo = request.form["titulo"]
        descricao = request.form["descricao"]
        horas_necessarias = request.form.get("horasNecessarias")
        data_prazo_conclusao = request.form.get("dataPrazoConclusao")
        valor_atividade = request.form.get("valorAtividade")
        prioridade_id = request.form["prioridadeId"]
        situacao_id = request.form["situacaoId"]
        supervisor_id = request.form.get("supervisorId")
        desenvolvedor_id = request.form.get("desenvolvedorId")
        usuario_solicitante_id = request.form.get("usuarioSolicitanteId")
        categoria_id = request.form.get("categoria_id", "").strip() or None
        area_id = request.form.get("area_id", "").strip() or None
        tags_ids_str = _extrair_tag_ids_do_request(request.form, "tag_id")

        campos = {
            "projetoId": ["Projeto", projeto_id],
            "titulo": ["Título", titulo],
            "prioridadeId": ["Prioridade", prioridade_id],
            "situacaoId": ["Situação", situacao_id]
        }

        validacao_campos_obrigatorios = ValidaForms.campo_obrigatorio(campos)

        if "validado" not in validacao_campos_obrigatorios:
            gravar_banco = False
            flash(("Verifique os campos destacados em vermelho!", "warning"))


        tags_processadas = _processar_tags_selecionadas(
            ids_str=tags_ids_str,
            campos_erros=validacao_campos_erros,
            dados_corretos=dados_corretos,
            key_erro="tag_id"
        )
        if "tag_id" in validacao_campos_erros:
            gravar_banco = False
        

        # Validação de horas necessárias
        horas_processadas = 0.0
        if horas_necessarias:
            try:
                horas_processadas = float(horas_necessarias.replace(',', '.'))
                if horas_processadas < 0:
                    validacao_campos_erros["horasNecessarias"] = "Horas não podem ser negativas"
                    gravar_banco = False
            except ValueError:
                validacao_campos_erros["horasNecessarias"] = "Valor inválido para horas necessárias"
                gravar_banco = False

        # Validação de valor monetário
        valor_atividade_100 = 0
        if valor_atividade:
            try:
                valor_limpo = valor_atividade.replace('R$', '').replace('.', '').replace(',', '.').strip()
                valor_float = float(valor_limpo)
                valor_atividade_100 = int(valor_float * 100)
                if valor_atividade_100 < 0:
                    validacao_campos_erros["valorAtividade"] = "Valor não pode ser negativo"
                    gravar_banco = False
            except ValueError:
                validacao_campos_erros["valorAtividade"] = "Valor monetário inválido"
                gravar_banco = False

        # Validação da data
        data_prazo = None
        if data_prazo_conclusao:
            try:
                data_prazo = datetime.strptime(data_prazo_conclusao, '%Y-%m-%d').date()
            except ValueError:
                validacao_campos_erros["dataPrazoConclusao"] = "Data inválida"
                gravar_banco = False

        # Processar anexos
        arquivos_anexos = request.files.getlist('anexos[]')
        anexos_validos = []
        
        for arquivo in arquivos_anexos:
            if arquivo and arquivo.filename:
                if not allowed_file(arquivo.filename):
                    validacao_campos_erros["anexos"] = f"Tipo de arquivo não permitido: {arquivo.filename}"
                    gravar_banco = False
                    continue
                
                # Validar tamanho
                arquivo.seek(0, os.SEEK_END)
                file_size = arquivo.tell()
                arquivo.seek(0)  # Reset para o início
                
                if file_size > MAX_FILE_SIZE:
                    validacao_campos_erros["anexos"] = f"Arquivo muito grande: {arquivo.filename} (máx. 10MB)"
                    gravar_banco = False
                    continue
                
                anexos_validos.append(arquivo)

        

        # Se passou em todas as validações, gravar no banco
        if gravar_banco:
            try:
                # Criar atividade
                atividade = AtividadeModel(
                    projeto_id=projeto_id,
                    titulo=titulo,
                    descricao=descricao,
                    supervisor_id=supervisor_id if supervisor_id else None,
                    desenvolvedor_id=desenvolvedor_id if desenvolvedor_id else None,
                    usuario_solicitante_id=usuario_solicitante_id if usuario_solicitante_id else None,
                    horas_necessarias=horas_processadas,
                    horas_utilizadas=0,
                    data_prazo_conclusao=data_prazo,
                    valor_atividade_100=valor_atividade_100,
                    prioridade_id=prioridade_id,
                    situacao_id=situacao_id,
                    categoria_id=categoria_id,
                    area_id=area_id 
                )
                
                db.session.add(atividade)
                db.session.flush()  # Para obter o ID da atividade

                atividade.tags = tags_processadas
                
                # Processar anexos usando a função upload_arquivo existente
                anexos_processados = 0
                for arquivo in anexos_validos:
                    try:
                        # Usar a função upload_arquivo existente
                        arquivo_model = upload_arquivo(
                            arquivo=arquivo,
                            pasta_destino='UPLOADED_ANEXOS_ATIVIDADES',
                            nome_referencia=f"ativ_{atividade.id}"
                        )
                        
                        if arquivo_model:
                            # Criar registro na tabela de anexos da atividade
                            anexo = AtividadeAnexoModel(
                                atividade_id=atividade.id,
                                nome_arquivo=arquivo_model.nome,
                                nome_original=arquivo.filename,
                                caminho_arquivo=arquivo_model.caminho,
                                tipo_arquivo=get_file_type(arquivo.filename),
                                tamanho=int(arquivo_model.tamanho),
                                mime_type=arquivo.content_type
                            )
                            db.session.add(anexo)
                            anexos_processados += 1
                        else:
                            flash((f"Erro ao fazer upload do arquivo: {arquivo.filename}", "warning"))
                            
                    except Exception as e:
                        flash((f"Erro ao processar anexo {arquivo.filename}: {str(e)}", "warning"))
                        continue
                
                # Se veio de uma solicitação aceita, marcar como processada
                if solicitacao_id and solicitacao:
                    try:
                        # Criar um relacionamento ou log da conversão se necessário
                        # Por enquanto, mantemos a solicitação como aceita
                        pass
                    except Exception as e:
                        # Se der erro aqui, não afeta o cadastro da atividade
                        flash((f"Aviso: Erro ao processar solicitação: {str(e)}", "warning"))
                
                db.session.commit()
                
                # Mensagem de sucesso com informações dos anexos
                if solicitacao_id:
                    mensagem_sucesso = f"Atividade criada com sucesso a partir da solicitação!"
                else:
                    mensagem_sucesso = "Atividade cadastrada com sucesso!"
                
                if anexos_processados > 0:
                    mensagem_sucesso += f" {anexos_processados} arquivo(s) anexado(s)."
                
                flash((mensagem_sucesso, "success"))
                return redirect(url_for("atividades_listar", projeto_id=projeto_id))
                
            except Exception as e:
                db.session.rollback()
                flash((f"Erro ao cadastrar atividade: {str(e)}", "warning"))
                gravar_banco = False

    return render_template(
        "sistema_wr/gerenciar/projetos/atividade_cadastrar.html",
        projetos=projetos,
        prioridades=prioridades,
        situacoes=situacoes,
        usuarios=usuarios,
        tags=tags,
        TagModel=TagModel,
        campos_obrigatorios=validacao_campos_obrigatorios,
        campos_erros=validacao_campos_erros,
        dados_corretos=dados_corretos,
        projeto_id_url=projeto_id,
        solicitacao_id=solicitacao_id,
        categorias=categorias,
        areas=areas
    )


@app.route("/gerenciar/atividades/editar/<int:atividade_id>", methods=["GET", "POST"])
@login_required
@requires_roles
def atividade_editar(atividade_id):
    """
    View para editar uma atividade existente
    """
    # Buscar a atividade
    atividade = AtividadeModel.obter_atividade_por_id(atividade_id)
    
    if not atividade:
        flash(("Atividade não encontrada!", "warning"))
        return redirect(url_for("atividades_listar"))
    
    #Opção de Marcar como concluído pelo usuário solicitante

    if request.method == 'POST':
        acao = request.form.get('acao')

        # Se a ação for de concluir
        if acao == 'concluir':
            # Confirmar que o usuário é o solicitante ou o supervisor
            if current_user.id == atividade.usuario_solicitante_id or current_user.id == atividade.supervisor_id:
                #Verifica se houve lançamento de horas
                from sistema.models_views.sistema_wr.gerenciar.projetos.lancamento_horas_model import LancamentoHorasModel
                
                total_lancamentos = LancamentoHorasModel.query.filter(
                    LancamentoHorasModel.atividade_id == atividade_id,
                    LancamentoHorasModel.deletado == False
                ).count()

                if total_lancamentos > 0:
                    try:
                        atividade.situacao_id = 4 #ID 4 = Concluída
                        atividade.data_alteracao = datetime.now()
                        db.session.commit()

                        # Enviar e-mail de conclusão
                        if atividade.usuario_solicitante and atividade.usuario_solicitante.email:
                            EmailAtividadeModel.enviar_email(
                                tipo_evento='atividade_concluida',
                                email_destino=atividade.usuario_solicitante.email,
                                solicitacao=None,
                                atividade=atividade,
                                usuario_acao=current_user
                            )

                        flash(('Atividade marcada como concluída!', 'success'))
                        return redirect(url_for('atividade_visualizar', atividade_id = atividade_id))
                    except Exception as e:
                        db.session.rollback()
                        flash((f'Erro ao concluir atividade: {str(e)}', 'error'))
                else:
                    flash(('A atividade deve ter pelo menos 1 lançamento de hora para ser concluída.', 'warning'))
            else:
                flash(('Apenas o solicitante ou supervisor pode concluir a atividade.', 'error'))
            
            # Redirecionar de volta para a edição se houve erro
            return redirect(url_for('atividade_editar', atividade_id=atividade_id))
        
    validacao_campos_obrigatorios = {}
    validacao_campos_erros = {}
    gravar_banco = True
    
    projetos = ProjetoModel.obter_projetos_asc_nome()
    prioridades = PrioridadeAtividadeModel.listar_prioridades_ativas()
    situacoes = AndamentoAtividadeModel.listar_andamentos_ativos()
    usuarios = UsuarioModel.obter_usuarios_asc_nome()
    tags = TagModel.listar_tags_ativas()
    categorias = CategoriaModel.listar_categorias_ativas()
    areas = AreaModel.listar_areas_ativas()
    

    if request.method == "POST":
        projeto_id = request.form.get("projetoId")
        titulo = request.form.get("titulo")
        descricao = request.form.get("descricao")
        usuario_execucao_id = request.form.get("usuarioExecucaoId")
        horas_necessarias = request.form.get("horasNecessarias")
        data_prazo_conclusao = request.form.get("dataPrazoConclusao")
        valor_atividade = request.form.get("valorAtividade")
        prioridade_id = request.form.get("prioridadeId")
        situacao_id = request.form.get("situacaoId")
        supervisor_id = request.form.get("supervisorId")
        desenvolvedor_id = request.form.get("desenvolvedorId")
        usuario_solicitante_id = request.form.get("usuarioSolicitanteId")
        categoria_id = request.form.get("categoria_id", "").strip() or None
        area_id = request.form.get("area_id", "").strip() or None
        
        tag_ids_str = _extrair_tag_ids_do_request(request.form, "tag_id")

        campos = {
            "projetoId": ["Projeto", projeto_id],
            "titulo": ["Título", titulo],
            "prioridadeId": ["Prioridade", prioridade_id],
            "situacaoId": ["Situação", situacao_id]
        }

        validacao_campos_obrigatorios = ValidaForms.campo_obrigatorio(campos)

        if "validado" not in validacao_campos_obrigatorios:
            gravar_banco = False
            flash(("Verifique os campos destacados em vermelho!", "warning"))

        tags_processadas = _processar_tags_selecionadas(
            ids_str=tag_ids_str,
            campos_erros=validacao_campos_erros,
            dados_corretos={},
            key_erro="tag_id"
        )
        if "tag_id" in validacao_campos_erros:
            gravar_banco = False

        horas_necessarias_processadas = 0.0
        if horas_necessarias:
            try:
                horas_necessarias_processadas = float(horas_necessarias.replace(',', '.'))
                if horas_necessarias_processadas < 0:
                    validacao_campos_erros["horasNecessarias"] = "Horas não podem ser negativas"
                    gravar_banco = False
            except ValueError:
                validacao_campos_erros["horasNecessarias"] = "Valor inválido para horas necessárias"
                gravar_banco = False

        valor_atividade_100 = 0
        if valor_atividade:
            try:
                valor_limpo = valor_atividade.replace('R$', '').replace('.', '').replace(',', '.').strip()
                valor_float = float(valor_limpo)
                valor_atividade_100 = int(valor_float * 100)
                if valor_atividade_100 < 0:
                    validacao_campos_erros["valorAtividade"] = "Valor não pode ser negativo"
                    gravar_banco = False
            except ValueError:
                validacao_campos_erros["valorAtividade"] = "Valor monetário inválido"
                gravar_banco = False

        # Validação da data
        data_prazo = None
        if data_prazo_conclusao:
            try:
                data_prazo = datetime.strptime(data_prazo_conclusao, '%Y-%m-%d').date()
            except ValueError:
                validacao_campos_erros["dataPrazoConclusao"] = "Data inválida"
                gravar_banco = False
                
        # Validação específica para conclusão de atividade
        if situacao_id == "4":  # 4 = Concluída
            from sistema.models_views.sistema_wr.gerenciar.projetos.lancamento_horas_model import LancamentoHorasModel
            
            # Verifica se existe pelo menos 1 lançamento de horas para esta atividade
            total_lancamentos = LancamentoHorasModel.query.filter(
                LancamentoHorasModel.atividade_id == atividade_id,
                LancamentoHorasModel.deletado == False
            ).count()
            
            if total_lancamentos == 0:
                flash((f"Atividade não pode ser concluída pois não tem registro de horas", "warning"))
                return redirect(url_for("atividade_editar", atividade_id=atividade_id))

        # Processa anexos (se houver novos uploads)
        arquivos_anexos = request.files.getlist('anexos[]')
        anexos_validos = []
        
        for arquivo in arquivos_anexos:
            if arquivo and arquivo.filename:
                if not allowed_file(arquivo.filename):
                    validacao_campos_erros["anexos"] = f"Tipo de arquivo não permitido: {arquivo.filename}"
                    gravar_banco = False
                    continue
                
                # Validar tamanho
                arquivo.seek(0, os.SEEK_END)
                file_size = arquivo.tell()
                arquivo.seek(0)  # Reset para o início
                
                if file_size > MAX_FILE_SIZE:
                    validacao_campos_erros["anexos"] = f"Arquivo muito grande: {arquivo.filename} (máx. 10MB)"
                    gravar_banco = False
                    continue
                
                anexos_validos.append(arquivo)
        if gravar_banco:
            try:
                # Atualizar dados da atividade
                atividade.projeto_id = projeto_id
                atividade.titulo = titulo
                atividade.descricao = descricao  # Mantém imagens Base64
                atividade.usuario_execucao_id = usuario_execucao_id if usuario_execucao_id else None
                atividade.horas_necessarias = horas_necessarias_processadas
                atividade.data_prazo_conclusao = data_prazo
                atividade.valor_atividade_100 = valor_atividade_100
                atividade.prioridade_id = prioridade_id
                atividade.situacao_id = situacao_id
                atividade.supervisor_id = supervisor_id if supervisor_id else None
                atividade.desenvolvedor_id = desenvolvedor_id if desenvolvedor_id else None
                atividade.usuario_solicitante_id = usuario_solicitante_id if usuario_solicitante_id else None
                atividade.categoria_id = categoria_id
                atividade.area_id = area_id
                atividade.tags = tags_processadas



                # Processa novos anexos (se houver)
                anexos_adicionados = 0
                for arquivo in anexos_validos:
                    try:
                        arquivo_model = upload_arquivo(
                            arquivo=arquivo,
                            pasta_destino='UPLOADED_ANEXOS_ATIVIDADES',
                            nome_referencia=f"ativ_{atividade.id}"
                        )
                        
                        if arquivo_model:
                            anexo = AtividadeAnexoModel(
                                atividade_id=atividade.id,
                                nome_arquivo=arquivo_model.nome,
                                nome_original=arquivo.filename,
                                caminho_arquivo=arquivo_model.caminho,
                                tipo_arquivo=get_file_type(arquivo.filename),
                                tamanho=int(arquivo_model.tamanho),
                                mime_type=arquivo.content_type
                            )
                            db.session.add(anexo)
                            anexos_adicionados += 1
                        else:
                            flash((f"Erro ao fazer upload do arquivo: {arquivo.filename}", "warning"))
                            
                    except Exception as e:
                        flash((f"Erro ao processar anexo {arquivo.filename}: {str(e)}", "warning"))
                        continue
                
                db.session.commit()

                # Enviar e-mail se a atividade foi marcada como concluída
                if situacao_id == "4" and atividade.usuario_solicitante and atividade.usuario_solicitante.email:
                    EmailAtividadeModel.enviar_email(
                        tipo_evento='atividade_concluida',
                        email_destino=atividade.usuario_solicitante.email, 
                        solicitacao=None,
                        atividade=atividade,
                        usuario_acao=current_user
                    )
                
                mensagem_sucesso = "Atividade atualizada com sucesso!"
                if anexos_adicionados > 0:
                    mensagem_sucesso += f" {anexos_adicionados} novo(s) anexo(s) adicionado(s)."
                
                flash((mensagem_sucesso, "success"))
                return redirect(url_for("atividade_visualizar", atividade_id=atividade.id))
                
            except Exception as e:
                db.session.rollback()
                flash((f"Erro ao atualizar atividade: {str(e)}", "error"))
                gravar_banco = False

    # Prepara dados corretos para preencher o formulário
    # Se houve erro, usa dados do formulário; senão, usa dados da atividade
    if request.method == "POST" and not gravar_banco:
        dados_corretos = request.form
    else:
        dados_corretos = {
            'projetoId': str(atividade.projeto_id),
            'titulo': atividade.titulo,
            'descricao': atividade.descricao or '',
            'horasNecessarias': str(atividade.horas_necessarias).replace('.', ','),
            'dataPrazoConclusao': atividade.data_prazo_conclusao.strftime('%Y-%m-%d') if atividade.data_prazo_conclusao else '',
            'valorAtividade': f"R$ {atividade.valor_atividade_100 / 100:.2f}".replace('.', ',') if atividade.valor_atividade_100 else 0,
            'prioridadeId': str(atividade.prioridade_id),
            'situacaoId': str(atividade.situacao_id),
            'supervisorId': str(atividade.supervisor_id) if atividade.supervisor_id else '',
            'desenvolvedorId': str(atividade.desenvolvedor_id) if atividade.desenvolvedor_id else '',
            'usuarioSolicitanteId': str(atividade.usuario_solicitante_id) if atividade.usuario_solicitante_id else '',
            'tag_id':[str(t.id) for t in (atividade.tags or [])],
            'categoria_id': str(atividade.categoria_id) if atividade.categoria_id else '',
            'area_id': str(atividade.area_id) if atividade.area_id else '',
        }

    return render_template(
        "sistema_wr/gerenciar/projetos/atividade_editar.html",
        atividade=atividade,
        projetos=projetos,
        prioridades=prioridades,
        situacoes=situacoes,
        usuarios=usuarios,
        tags=tags,
        categorias=categorias,
        areas=areas,
        TagModel=TagModel,
        campos_obrigatorios=validacao_campos_obrigatorios,
        campos_erros=validacao_campos_erros,
        dados_corretos=dados_corretos
    )
    
    
@app.route("/gerenciar/atividades/anexo/remover/<int:anexo_id>", methods=["POST"])
@login_required
@requires_roles
def atividade_remover_anexo(anexo_id):
    """
    View para remover anexo da atividade via AJAX
    """
    try:
        anexo = AtividadeAnexoModel.query.filter(
            AtividadeAnexoModel.id == anexo_id,
            AtividadeAnexoModel.deletado == False
        ).first()
        
        if not anexo:
            return jsonify({
                'success': False,
                'message': 'Anexo não encontrado!'
            }), 404
        
        anexo.deletado = True
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Anexo removido com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Erro ao remover anexo: {str(e)}'
        }), 500


@app.route("/gerenciar/atividades/excluir/<int:id>", methods=["GET", "POST"])
@login_required
@requires_roles
def atividade_excluir(id):
    atividade = AtividadeModel.obter_atividade_por_id(id)
    
    if atividade:
        if atividade.situacao_id != 1:
            flash((f"Somente atividades 'Não iniciadas' podem ser excluidas!", "warning"))
            return redirect(url_for("atividades_listar"))
            
        atividade.deletado = True
        db.session.commit()
        
        flash((f"Atividade excluida com sucesso!", "success"))
        return redirect(url_for("atividades_listar"))
    
    else:
        flash((f"Atividade não encontrada!!", "warning"))
        return redirect(url_for("atividades_listar"))


@app.route("/gerenciar/atividades/duplicar/<int:id>", methods=["POST"])
@login_required
@requires_roles
def atividade_duplicar(id):
    atividade_original = AtividadeModel.obter_atividade_por_id(id)
    
    if not atividade_original:
        return jsonify({'success': False, 'message': 'Atividade não encontrada'}), 404
    
    try:
        # Criar nova atividade com dados da original
        nova_atividade = AtividadeModel(
            projeto_id=atividade_original.projeto_id,
            titulo=f"{atividade_original.titulo} (Cópia)",
            descricao=atividade_original.descricao,
            supervisor_id=atividade_original.supervisor_id,  # Corrigido
            desenvolvedor_id=atividade_original.desenvolvedor_id,  # Corrigido
            usuario_solicitante_id=atividade_original.usuario_solicitante_id,  # Corrigido
            horas_necessarias=atividade_original.horas_necessarias,
            horas_utilizadas=0,  # Zerar horas utilizadas
            data_prazo_conclusao=atividade_original.data_prazo_conclusao,
            valor_atividade_100=atividade_original.valor_atividade_100,
            prioridade_id=atividade_original.prioridade_id,
            situacao_id=atividade_original.situacao_id
        )
        
        db.session.add(nova_atividade)
        db.session.flush()  # Para obter o ID da nova atividade

        nova_atividade.tags = list(atividade_original.tags or [])

        db.session.commit()
        return jsonify({
            'success': True, 
            'message': 'Atividade duplicada com sucesso!',
            'nova_atividade_id': nova_atividade.id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False, 
            'message': f'Erro ao duplicar atividade: {str(e)}'
        }), 500